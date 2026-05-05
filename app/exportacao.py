# -*- coding: utf-8 -*-
"""
Exportação Excel com openpyxl.
"""
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_COR = {
    "OK":           "D4EDDA",
    "Inconsistente": "F8D7DA",
    "header_aud":   "1565C0",
    "header_inc":   "0D6EFD",
    "header_just":  "6F42C1",
}

_COLUNAS_AUDITORIA = [
    "Funcionário", "Departamento", "Cod. Funcionário",
    "Tem Direito", "Está na Fatura", "Está na Compra",
    "Valor Contrato", "Valor Empresa (Compra)", "Valor Fatura", "Valor Compra Total",
    "Dif. Contrato x Compra", "Dif. Fatura x Compra",
    "Data Inclusão", "Dt. Admissão", "Dt. Demissão",
    "Status", "Inconsistência", "Ação Sugerida",
]

_COLUNAS_INCLUSOES = [
    "Funcionário", "Departamento", "Dt. Admissão", "Data Inclusão",
    "Período", "Valor Fatura", "Valor Contrato", "Tem Direito", "Contrato",
]

_MOEDA = {"Valor Fatura", "Valor Empresa (Compra)", "Valor Compra Total",
          "Valor Contrato", "Dif. Contrato x Compra", "Dif. Fatura x Compra"}


def _header(ws, colunas: list, cor_hex: str):
    fill = PatternFill("solid", fgColor=cor_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    for i, col in enumerate(colunas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _ajustar_larguras(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 55)


def _escrever_dados(ws, df: pd.DataFrame, colunas: list, cor_status=True):
    for row_i, (_, row) in enumerate(df.iterrows(), 2):
        status = str(row.get("Status", "") or "")
        fill = PatternFill("solid", fgColor=_COR.get(status, "FFFFFF")) if cor_status else None
        for col_i, col in enumerate(colunas, 1):
            val = row.get(col, "")
            if col in _MOEDA:
                try:
                    val = float(val)
                except Exception:
                    pass
            c = ws.cell(row=row_i, column=col_i, value=val)
            if fill:
                c.fill = fill
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)


def exportar_excel(df: pd.DataFrame,
                   df_inc: pd.DataFrame,
                   df_just: pd.DataFrame,
                   periodo: str) -> bytes:
    wb = Workbook()

    # ── Aba 1: Auditoria ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Auditoria"
    cols1 = [c for c in _COLUNAS_AUDITORIA if c in df.columns]
    _header(ws1, cols1, _COR["header_aud"])
    if not df.empty:
        _escrever_dados(ws1, df, cols1)
    _ajustar_larguras(ws1)
    ws1.freeze_panes = "A2"

    # ── Aba 2: Inclusões do Mês ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Inclusões do Mês")
    cols2 = [c for c in _COLUNAS_INCLUSOES if c in df_inc.columns] if not df_inc.empty else []
    if cols2:
        _header(ws2, cols2, _COR["header_inc"])
        _escrever_dados(ws2, df_inc, cols2, cor_status=False)
        _ajustar_larguras(ws2)
        ws2.freeze_panes = "A2"
    else:
        ws2.cell(row=1, column=1, value=f"Nenhuma inclusão no mês {periodo}.")

    # ── Aba 3: Justificativas ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Justificativas")
    if not df_just.empty:
        cols3 = [c for c in ["Funcionário", "Inconsistência", "Justificativa", "Criado em"]
                 if c in df_just.columns]
        if cols3:
            _header(ws3, cols3, _COR["header_just"])
            _escrever_dados(ws3, df_just, cols3, cor_status=False)
            _ajustar_larguras(ws3)
    else:
        ws3.cell(row=1, column=1, value="Nenhuma justificativa registrada.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
